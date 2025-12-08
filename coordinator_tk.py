import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import requests
import socketio
import threading
from openpyxl import load_workbook, Workbook

# Server URLs
SERVER_UPLOAD = "http://192.168.1.19:5003/upload_matrix"
SERVER_WS = "http://192.168.1.19:5003"


class CoordinatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("DCTW Coordinator")
        self.root.geometry("900x600")

        self.matrix = []
        self.entries = []
        self.deciders_local = [
            {"name": "decider_policeman", "weight": 40.0},
            {"name": "decider_economist", "weight": 25.0},
            {"name": "decider_environmental representative", "weight": 20.0},
            {"name": "decider_public representative", "weight": 15.0},
        ]

        self.sio = None
        self.sio_thread = None
        self.received_rankings = {}
        
        # Négociation variables
        self.negotiation_window = None
        self.negotiation_log = None
        self.current_action_proposal = None
        self.next_action_suggestion = None  # Suggestion automatique pour la prochaine action
        self.negotiation_responses = {}
        self.action_scores = {}  # Stockage des scores d'actions
        self.sorted_actions = []  # Actions triées par score
        self.best_action = None  # Meilleure action basée sur scoring
        self.progress_label = None  # Label pour la progression
        self.next_action_label = None  # Label pour la suggestion suivante

        # Top frame
        top = ttk.Frame(root)
        top.pack(fill="x", padx=10, pady=8)
        self.info_label = ttk.Label(top, text="🔌 Connecting to server...", font=("Arial", 14, "bold"))
        self.info_label.pack(side="left")

        # Buttons
        btn_frame = ttk.Frame(root)
        btn_frame.pack(fill="x", padx=10, pady=6)
        ttk.Button(btn_frame, text="📂 Upload Excel", command=self.load_excel).pack(side="left", padx=4)
        ttk.Button(btn_frame, text="➕ New", command=self.create_matrix_dialog).pack(side="left", padx=4)
        self.save_btn = ttk.Button(btn_frame, text="💾 Save Excel", command=self.save_excel, state="disabled")
        self.save_btn.pack(side="left", padx=4)
        self.send_btn = ttk.Button(btn_frame, text="🚀 Send Matrix", command=self.send_matrix, state="disabled")
        self.send_btn.pack(side="left", padx=4)
        ttk.Button(btn_frame, text="👥 Show Deciders", command=self.show_deciders_local).pack(side="left", padx=4)
        self.aggregate_btn = ttk.Button(btn_frame, text="⚙️ Aggregate",
                                        command=self.aggregate_action,
                                        state="disabled")
        self.aggregate_btn.pack(side="left", padx=4)


        # Rankings frame (au-dessus de la matrice)
        self.rankings_frame = ttk.Frame(root)
        self.rankings_frame.pack(fill="x", padx=10, pady=6)

        # Canvas pour la matrice scrollable
        self.canvas = tk.Canvas(root)
        self.frame_container = ttk.Frame(self.canvas)
        self.vsb = ttk.Scrollbar(root, orient="vertical", command=self.canvas.yview)
        self.hsb = ttk.Scrollbar(root, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=self.vsb.set, xscrollcommand=self.hsb.set)
        self.vsb.pack(side="right", fill="y")
        self.hsb.pack(side="bottom", fill="x")
        self.canvas.pack(fill="both", expand=True, padx=10, pady=6)
        self.canvas.create_window((0, 0), window=self.frame_container, anchor="nw")
        self.frame_container.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))

        self.start_socketio_client()

    # ------------------- SOCKET.IO -------------------
    def start_socketio_client(self):
        def run_client():
            try:
                self.sio = socketio.Client(logger=True, reconnection=True)

                @self.sio.on('final_ranking')
                def on_final_ranking(data):
                    decider_name = data['decider']
                    print(f"📊 Received ranking from {decider_name}")
                    
                    self.received_rankings[decider_name] = {
                        'phi': data.get('phi', []),
                        'ranking': data['ranking']
                    }
                    
                    # Mettre à jour l'interface
                    self.root.after(0, self._update_status)
                    
                    # Activer les boutons si tous les classements sont reçus
                    if len(self.received_rankings) == 4:
                        self.root.after(0, lambda: self.aggregate_btn.config(state="normal"))
                        self.root.after(0, lambda: self.negotiation_btn.config(state="normal"))
                        self.root.after(0, lambda: self.info_label.config(
                            text=f"✅ All rankings received - Ready for negotiation"
                        ))

                @self.sio.on("negotiation_response")
                def on_negotiation_response(data):
                    decider = data["decider"]
                    answer = data["answer"]
                    action = data["action"]
                    
                    print(f"📩 Response from {decider}: {answer}")
                    
                    # Stocker la réponse
                    self.negotiation_responses[decider] = answer
                    
                    # Mettre à jour le journal
                    if self.negotiation_log:
                        self.root.after(0, lambda: self._update_log_response(decider, answer))
                    
                    # Vérifier si tous les décideurs ont répondu
                    if len(self.negotiation_responses) == 4:
                        accept_count = sum(1 for ans in self.negotiation_responses.values() if ans == "accept")
                        accept_ratio = accept_count / 4
                        
                        # Mettre à jour le journal avec le résultat
                        if self.negotiation_log:
                            self.root.after(0, lambda: self._update_log_result(action, accept_ratio))
                        
                        if accept_ratio >= 0.9:  # 90% threshold
                            # Notifier tous les décideurs
                            self.sio.emit("negotiation_selected", {"action": action})
                            
                            if self.negotiation_log:
                                self.root.after(0, lambda: self._update_log_selected(action, accept_ratio))
                            
                            self.root.after(0, lambda: messagebox.showinfo(
                                "Negotiation Finished", 
                                f"Action '{action}' has been selected with {accept_ratio:.0%} acceptance!"
                            ))
                            
                            # Désactiver le bouton d'envoi
                            if hasattr(self, 'send_action_btn'):
                                self.root.after(0, lambda: self.send_action_btn.config(state="disabled"))
                            
                            # Effacer la suggestion suivante
                            self.next_action_suggestion = None
                            if self.next_action_label:
                                self.root.after(0, lambda: self.next_action_label.config(text="No next action suggested"))
                        else:
                            if self.negotiation_log:
                                self.root.after(0, lambda: self._update_log_rejected(action, accept_ratio))
                            
                            # Calculer automatiquement la prochaine action suggérée
                            self.root.after(500, self._suggest_next_action)
                            
                            # Réactiver le bouton pour que le coordinateur puisse envoyer manuellement
                            if hasattr(self, 'send_action_btn'):
                                self.root.after(0, lambda: self.send_action_btn.config(state="normal"))

                @self.sio.on("negotiation_selected")
                def on_selected(data):
                    action = data["action"]
                    print(f"✅ Final selection: {action}")
                    
                    if self.negotiation_log:
                        self.root.after(0, lambda: self.negotiation_log.insert(
                            "end", f"\n🎯 Server confirms: Action '{action}' selected!\n"
                        ))
                        self.root.after(0, lambda: self.negotiation_log.see("end"))

                @self.sio.event
                def connect():
                    print("✅ Coordinator connected to server")
                    self.root.after(0, lambda: self.info_label.config(
                        text="✅ Connected to server - Ready"
                    ))

                @self.sio.event
                def disconnect():
                    print("⚠️ Coordinator disconnected from server")
                    self.root.after(0, lambda: self.info_label.config(
                        text="🔌 Disconnected from server"
                    ))

                # Se connecter au serveur
                print("🔗 Connecting to server...")
                self.sio.connect(SERVER_WS)
                print("✅ Socket.IO connection established")
                self.sio.wait()
                
            except Exception as e:
                print(f"❌ SocketIO Error: {e}")
                self.root.after(0, lambda: self.info_label.config(
                    text=f"❌ Connection error: {str(e)[:50]}"
                ))

        self.sio_thread = threading.Thread(target=run_client, daemon=True)
        self.sio_thread.start()

    def _update_log_response(self, decider, answer):
        """Mettre à jour le journal avec une réponse"""
        if self.negotiation_log:
            self.negotiation_log.config(state="normal")
            self.negotiation_log.insert("end", f"   {decider} → {answer}\n")
            self.negotiation_log.see("end")
            self.negotiation_log.config(state="disabled")

    def _update_log_result(self, action, accept_ratio):
        """Mettre à jour le journal avec le résultat"""
        if self.negotiation_log:
            self.negotiation_log.config(state="normal")
            self.negotiation_log.insert("end", f"\n📊 Results for '{action}':\n")
            self.negotiation_log.insert("end", f"   Total 'yes' = {accept_ratio:.0%}\n")
            
            if accept_ratio >= 0.9:
                self.negotiation_log.insert("end", f"   ✅ ≥90% - Action SELECTED!\n")
            else:
                self.negotiation_log.insert("end", f"   ❌ <90% - Action REJECTED\n")
            
            self.negotiation_log.see("end")
            self.negotiation_log.config(state="disabled")

    def _update_log_selected(self, action, accept_ratio):
        """Mettre à jour le journal avec la sélection"""
        if self.negotiation_log:
            self.negotiation_log.config(state="normal")
            self.negotiation_log.insert("end", f"\n🎉 Action SELECTED! ({accept_ratio:.0%} accept)\n")
            self.negotiation_log.insert("end", f"✅ Action '{action}' has been selected!\n")
            self.negotiation_log.see("end")
            self.negotiation_log.config(state="disabled")

    def _update_log_rejected(self, action, accept_ratio):
        """Mettre à jour le journal avec le rejet"""
        if self.negotiation_log:
            self.negotiation_log.config(state="normal")
            self.negotiation_log.insert("end", f"\n❌ Action REJECTED ({accept_ratio:.0%} acceptance)\n")
            self.negotiation_log.insert("end", f"Next action suggested automatically. Click 'Send Action' to try it.\n")
            self.negotiation_log.see("end")
            self.negotiation_log.config(state="disabled")

    # ------------------- SUGGESTION AUTOMATIQUE -------------------
    def _suggest_next_action(self):
        """Suggérer automatiquement l'action suivante"""
        if not self.sorted_actions:
            if self.negotiation_log:
                self.negotiation_log.config(state="normal")
                self.negotiation_log.insert("end", "\n❌ No more actions to suggest!\n")
                self.negotiation_log.see("end")
                self.negotiation_log.config(state="disabled")
            return
        
        # Trouver l'action actuelle dans la liste
        current_idx = -1
        for i, (action, score) in enumerate(self.sorted_actions):
            if action == self.current_action_proposal:
                current_idx = i
                break
        
        # Prendre la suivante
        next_idx = current_idx + 1
        if next_idx < len(self.sorted_actions):
            next_action, next_score = self.sorted_actions[next_idx]
            
            # Stocker la suggestion
            self.next_action_suggestion = next_action
            
            # Mettre à jour l'indicateur de progression
            self.root.after(0, self._update_progress_label)
            
            # Mettre à jour le label de suggestion
            if self.next_action_label:
                self.root.after(0, lambda: self.next_action_label.config(
                    text=f"Suggested next action (#{next_idx+1}): '{next_action}' (Score: {next_score:.2f})",
                    foreground="white"
                ))
            
            # Mettre à jour le journal
            if self.negotiation_log:
                self.negotiation_log.config(state="normal")
                self.negotiation_log.insert("end", f"\n💡 NEXT ACTION SUGGESTED\n")
                self.negotiation_log.insert("end", f"Suggested action (#{next_idx+1}): '{next_action}'\n")
                self.negotiation_log.insert("end", f"Score: {next_score:.2f}\n")
                self.negotiation_log.insert("end", f"Click 'Send Action to Deciders' to try it.\n")
                self.negotiation_log.see("end")
                self.negotiation_log.config(state="disabled")
            
            # Mettre à jour le texte du bouton
            if hasattr(self, 'send_action_btn'):
                self.root.after(0, lambda: self.send_action_btn.config(
                    text=f"📨 Send '{next_action}' to Deciders"
                ))
            
            print(f"💡 Suggested next action: {next_action} (rank #{next_idx+1})")
        else:
            if self.negotiation_log:
                self.negotiation_log.config(state="normal")
                self.negotiation_log.insert("end", "\n❌ No more actions to suggest!\n")
                self.negotiation_log.see("end")
                self.negotiation_log.config(state="disabled")
            
            if self.next_action_label:
                self.root.after(0, lambda: self.next_action_label.config(
                    text="No more actions available!",
                    foreground="red"
                ))

    def _update_progress_label(self):
        """Mettre à jour l'indicateur de progression"""
        if hasattr(self, 'progress_label') and self.current_action_proposal and self.sorted_actions:
            # Trouver l'index de l'action actuelle
            for idx, (action, score) in enumerate(self.sorted_actions):
                if action == self.current_action_proposal:
                    current_try = idx + 1
                    total_actions = len(self.sorted_actions)
                    
                    # Mettre à jour le label
                    self.progress_label.config(text=f"Current: #{current_try} of {total_actions} total")
                    
                    # Changer la couleur selon la position
                    if current_try == 1:
                        self.progress_label.config(foreground="green")
                    elif current_try <= 5:
                        self.progress_label.config(foreground="orange")
                    else:
                        self.progress_label.config(foreground="red")
                    break

    # ------------------- NÉGOCIATION PANEL -------------------
    def open_negotiation_panel(self):
        """Ouvre le panneau de négociation"""
        if self.negotiation_window and self.negotiation_window.winfo_exists():
            self.negotiation_window.lift()
            return
        
        self.negotiation_window = tk.Toplevel(self.root)
        self.negotiation_window.title("Negotiation Panel")
        self.negotiation_window.geometry("700x600")
        
        # Titre
        ttk.Label(self.negotiation_window, text="🤝 NEGOTIATION PANEL", 
                 font=("Arial", 14, "bold")).pack(pady=10)
        
        # Calculer les scores si nécessaire
        if not self.action_scores and self.matrix:
            self._calculate_action_scores()
        
        # Déterminer la meilleure action
        if self.action_scores:
            self.sorted_actions = sorted(self.action_scores.items(), key=lambda x: x[1], reverse=True)
            self.best_action, best_score = self.sorted_actions[0]
        
        # Frame pour l'action actuelle
        current_frame = ttk.LabelFrame(self.negotiation_window, text="Current Action", padding=10)
        current_frame.pack(fill="x", padx=10, pady=5)
        
        if self.best_action:
            best_score = self.action_scores[self.best_action]
            current_rank = 1
            
            ttk.Label(current_frame, 
                     text=f"Action to send: '{self.best_action}'", 
                     font=("Arial", 11, "bold")).pack(anchor="w")
            ttk.Label(current_frame, 
                     text=f"Score: {best_score:.2f} | Rank: #{current_rank}",
                     font=("Arial", 10)).pack(anchor="w", pady=(2, 5))
        else:
            ttk.Label(current_frame, 
                     text="No action scores calculated yet.", 
                     font=("Arial", 10)).pack()
        
        # Frame pour la suggestion suivante
        suggestion_frame = ttk.LabelFrame(self.negotiation_window, text="Next Action Suggestion", padding=10)
        suggestion_frame.pack(fill="x", padx=10, pady=5)
        
        self.next_action_label = ttk.Label(suggestion_frame, 
                                         text="No suggestion yet. Send first action to start.",
                                         font=("Arial", 10),
                                         foreground="gray")
        self.next_action_label.pack(anchor="w")
        
        # Indicateur de progression
        progress_frame = ttk.Frame(self.negotiation_window)
        progress_frame.pack(fill="x", padx=10, pady=5)
        
        self.progress_label = ttk.Label(progress_frame, 
                                       text="Progress: Not started",
                                       font=("Arial", 9, "bold"),
                                       foreground="gray")
        self.progress_label.pack(side="left")
        
        # Bouton pour envoyer l'action
        btn_frame = ttk.Frame(self.negotiation_window)
        btn_frame.pack(fill="x", padx=10, pady=10)
        
        self.send_action_btn = ttk.Button(btn_frame, text="📨 Send Action to Deciders", 
                                         command=self.send_current_action)
        self.send_action_btn.pack()
        
        # Frame pour toutes les actions (pour référence)
        all_actions_frame = ttk.LabelFrame(self.negotiation_window, text="All Actions (by score)", padding=10)
        all_actions_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        # Liste de toutes les actions triées
        list_frame = ttk.Frame(all_actions_frame)
        list_frame.pack(fill="both", expand=True)
        
        self.actions_listbox = tk.Listbox(list_frame, height=8, font=("Arial", 10))
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.actions_listbox.yview)
        self.actions_listbox.configure(yscrollcommand=scrollbar.set)
        
        self.actions_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Remplir avec les actions triées par score
        if self.sorted_actions:
            for i, (action, score) in enumerate(self.sorted_actions, 1):
                prefix = "➤ " if action == self.best_action else "  "
                self.actions_listbox.insert("end", f"{prefix}{i}. {action} (score: {score:.2f})")
        
        # Journal de négociation
        ttk.Label(self.negotiation_window, text="Negotiation Log:", 
                 font=("Arial", 10, "bold")).pack(anchor="w", padx=10, pady=(10, 0))
        
        log_frame = ttk.Frame(self.negotiation_window)
        log_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.negotiation_log = tk.Text(log_frame, height=12, wrap="word", font=("Arial", 9))
        log_scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.negotiation_log.yview)
        self.negotiation_log.configure(yscrollcommand=log_scrollbar.set)
        
        self.negotiation_log.pack(side="left", fill="both", expand=True)
        log_scrollbar.pack(side="right", fill="y")
        
        # Message initial
        self.negotiation_log.insert("end", "📋 NEGOTIATION PANEL READY\n")
        self.negotiation_log.insert("end", "="*50 + "\n")
        if self.best_action:
            self.negotiation_log.insert("end", f"Best action: {self.best_action} (selected from scoring)\n")
            self.negotiation_log.insert("end", "Click 'Send Action to Deciders' to start negotiation.\n")
        else:
            self.negotiation_log.insert("end", "No action available for negotiation.\n")
        self.negotiation_log.insert("end", "="*50 + "\n\n")
        self.negotiation_log.config(state="disabled")
        
        # Bouton pour fermer
        ttk.Button(self.negotiation_window, text="Close", 
                  command=self.negotiation_window.destroy).pack(pady=10)
    
    def _calculate_action_scores(self):
        """Calculer les scores des actions"""
        if not self.matrix or len(self.matrix) < 2:
            return
        
        actions = [row[0] for row in self.matrix[1:]]
        n_actions = len(actions)
        
        # Réinitialiser les scores
        self.action_scores = {action: 0.0 for action in actions}
        
        # Calculer les scores pondérés
        for decider in self.deciders_local:
            name = decider["name"]
            weight = decider["weight"] / 100.0
            ranking_data = self.received_rankings.get(name)
            
            if not ranking_data:
                continue
                
            ranking = ranking_data["ranking"]
            
            for pos, action_idx in enumerate(ranking):
                if action_idx < len(actions):
                    action_name = actions[action_idx]
                    score = (n_actions - pos) * weight
                    self.action_scores[action_name] += score
    
    def send_current_action(self):
        """Envoyer l'action actuelle aux décideurs"""
        # Si une suggestion est disponible, l'utiliser
        if self.next_action_suggestion:
            action_to_send = self.next_action_suggestion
        elif self.best_action:
            action_to_send = self.best_action
        else:
            messagebox.showwarning("No Action", "No action determined yet.")
            return
        
        # Définir l'action actuelle
        self.current_action_proposal = action_to_send
        
        # Réinitialiser les réponses
        self.negotiation_responses = {}
        
        # Mettre à jour l'indicateur de progression
        self.root.after(0, self._update_progress_label)
        
        # Mettre à jour le journal
        if self.negotiation_log:
            # Trouver le rang de l'action
            action_rank = 1
            action_score = 0
            for idx, (action, score) in enumerate(self.sorted_actions):
                if action == action_to_send:
                    action_rank = idx + 1
                    action_score = score
                    break
            
            self.negotiation_log.config(state="normal")
            self.negotiation_log.insert("end", f"\n📤 SENDING ACTION #{action_rank}\n")
            self.negotiation_log.insert("end", f"Action: '{action_to_send}'\n")
            self.negotiation_log.insert("end", f"Score: {action_score:.2f} | Rank: #{action_rank}\n")
            self.negotiation_log.insert("end", f"Deciders responses:\n")
            self.negotiation_log.see("end")
            self.negotiation_log.config(state="disabled")
        
        # Désactiver le bouton pendant l'attente
        if hasattr(self, 'send_action_btn'):
            self.send_action_btn.config(state="disabled")
        
        # Envoyer au serveur
        try:
            self.sio.emit("negotiation_proposal", {"action": action_to_send})
            print(f"📨 Sent proposal for action: {action_to_send}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to send proposal: {e}")
            if hasattr(self, 'send_action_btn'):
                self.send_action_btn.config(state="normal")
    
    

    # ------------------- STATUS & RANKINGS -------------------
    def _update_status(self):
        received_count = len(self.received_rankings)
        total_deciders = len(self.deciders_local)
        status_text = f"📊 {received_count}/{total_deciders} rankings received"
        self.info_label.config(text=status_text)
        self.display_rankings_above_matrix()

    def display_rankings_above_matrix(self):
        # Clear previous
        for widget in self.rankings_frame.winfo_children():
            widget.destroy()

        if not self.received_rankings:
            return

        actions = [row[0] for row in self.matrix[1:]] if self.matrix and len(self.matrix) > 1 else []

        for decider_name, data in self.received_rankings.items():
            ttk.Label(self.rankings_frame, text=f"{decider_name} Ranking:", font=("Arial", 10, "bold")).pack(anchor="w", padx=5, pady=(5,2))

            tree = ttk.Treeview(self.rankings_frame, columns=("Rank", "Action"), show="headings", height=4)
            tree.heading("Rank", text="#")
            tree.heading("Action", text="Action")
            tree.column("Rank", width=50, anchor="center")
            tree.column("Action", width=300)
            tree.pack(fill="x", padx=5, pady=2)

            for pos, action_idx in enumerate(data['ranking']):
                action_name = actions[action_idx] if action_idx < len(actions) else f"Action {action_idx+1}"
                tree.insert("", "end", values=(pos+1, action_name))

    def show_decider_ranking(self, parent_win, decider_name):
        # Clear previous tree if exists
        for widget in parent_win.winfo_children():
            if isinstance(widget, ttk.Treeview):
                widget.destroy()

        actions = [row[0] for row in self.matrix[1:]] if self.matrix and len(self.matrix) > 1 else []
        data = self.received_rankings[decider_name]

        ttk.Label(parent_win, text=f"{decider_name} Ranking:", font=("Arial", 10, "bold")).pack(anchor="w", padx=10, pady=(10,2))

        tree = ttk.Treeview(parent_win, columns=("Rank", "Action"), show="headings", height=35)
        tree.heading("Rank", text="#")
        tree.heading("Action", text="Action")
        tree.column("Rank" )
        tree.column("Action")
        tree.pack(fill="both", expand=False, padx=5, pady=3)

        for pos, action_idx in enumerate(data['ranking']):
            action_name = actions[action_idx] if action_idx < len(actions) else f"Action {action_idx+1}"
            tree.insert("", "end", values=(pos+1, action_name))

    # ------------------- AGGREGATION -------------------
    def aggregate_action(self):
        win = tk.Toplevel(self.root)
        win.title("Calculate Action Scores")
        win.geometry("300x150")
        win.resizable(False, False)

        ttk.Label(win, text="Calculate Action Scores", 
                 font=("Arial", 12, "bold")).pack(pady=15)

        score_btn = ttk.Button(win, text="⚙️ Calculate Scores", 
                              command=self.open_scoring_window)
        score_btn.pack(pady=10)

    def open_scoring_window(self):
        # Vérifier que tous les classements sont reçus
        if len(self.received_rankings) < len(self.deciders_local):
            messagebox.showwarning("Warning", "Not all decider rankings received yet.")
            return

        # Calculer les scores
        self._calculate_action_scores()
        
        # Déterminer la meilleure action
        if self.action_scores:
            self.sorted_actions = sorted(self.action_scores.items(), key=lambda x: x[1], reverse=True)
            self.best_action, best_score = self.sorted_actions[0]
        
        # Créer la fenêtre de scoring
        win = tk.Toplevel(self.root)
        win.title("Action Scoring Results")
        win.geometry("600x500")

        ttk.Label(win, text="📊 ACTION SCORING RESULTS", 
                 font=("Arial", 14, "bold")).pack(pady=10)

        # Frame principal
        main_frame = ttk.Frame(win)
        main_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Tableau des scores
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill="both", expand=True)

        tree = ttk.Treeview(tree_frame, columns=("Rank", "Action", "Score"), 
                           show="headings", height=15)
        
        # Configurer les colonnes
        tree.heading("Rank", text="Rank")
        tree.heading("Action", text="Action")
        tree.heading("Score", text="Score")
        
        tree.column("Rank", width=50, anchor="center")
        tree.column("Action", width=300, anchor="w")
        tree.column("Score", width=100, anchor="center")
        
        # Barre de défilement
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Remplir le tableau avec les actions triées par score
        for rank, (action, score) in enumerate(self.sorted_actions, 1):
            tree.insert("", "end", values=(rank, action, f"{score:.2f}"))

        # Boutons en bas
        btn_frame = ttk.Frame(win)
        btn_frame.pack(fill="x", pady=10, padx=20)
        
        # Bouton pour ouvrir le panneau de négociation
        def open_negotiation():
            win.destroy()
            self.open_negotiation_panel()
        
        start_btn = ttk.Button(btn_frame, text="🤝 Open Negotiation Panel",
                              command=open_negotiation)
        start_btn.pack(side="left", padx=5)
        
        ttk.Button(btn_frame, text="Close", 
                  command=win.destroy).pack(side="right", padx=5)

    # ------------------- MATRIX GRID -------------------
    def clear_grid(self):
        for row in self.entries:
            for e in row:
                e.destroy()
        self.entries = []

    def build_grid(self):
        self.clear_grid()
        rows = len(self.matrix)
        cols = len(self.matrix[0]) if rows > 0 else 0
        self.entries = [[None for _ in range(cols)] for _ in range(rows)]
        
        for i in range(rows):
            for j in range(cols):
                val = self.matrix[i][j] if i < len(self.matrix) and j < len(self.matrix[i]) else ""
                e = ttk.Entry(self.frame_container, width=15)
                e.grid(row=i, column=j, padx=2, pady=2, sticky="nsew")
                e.insert(0, val)
                e.bind("<KeyRelease>", lambda ev: self.on_edit())
                self.entries[i][j] = e

    def update_matrix_from_entries(self):
        self.matrix = [[cell.get() for cell in row] for row in self.entries]

    def on_edit(self):
        self.save_btn.config(state="normal")
        self.send_btn.config(state="normal")

    # ------------------- EXCEL -------------------
    def load_excel(self):
        path = filedialog.askopenfilename(title="Select Excel file", 
                                         filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        
        try:
            wb = load_workbook(path)
            ws = wb.active
            self.matrix = [[str(c) if c is not None else "" for c in r] 
                          for r in ws.iter_rows(values_only=True)]
            
            if not self.matrix:
                messagebox.showwarning("Empty", "The Excel sheet is empty.")
                return
            
            self.build_grid()
            self.save_btn.config(state="normal")
            self.send_btn.config(state="normal")
            self.info_label.config(text=f"📂 Loaded: {path.split('/')[-1]}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Cannot load Excel file: {e}")

    def create_matrix_dialog(self):
        rows = simpledialog.askinteger("Rows", "How many choices (rows)?", 
                                      initialvalue=3, minvalue=1, maxvalue=50)
        if rows is None:
            return
            
        cols = simpledialog.askinteger("Columns", "How many criteria (columns)?", 
                                      initialvalue=3, minvalue=1, maxvalue=50)
        if cols is None:
            return
            
        self.matrix = [["" for _ in range(cols)] for _ in range(rows)]
        self.build_grid()
        self.save_btn.config(state="normal")
        self.send_btn.config(state="normal")
        self.info_label.config(text="➕ New Matrix Created")

    def save_excel(self):
        self.update_matrix_from_entries()
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                           filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
            
        try:
            wb = Workbook()
            ws = wb.active
            for i, row in enumerate(self.matrix, start=1):
                for j, val in enumerate(row, start=1):
                    ws.cell(row=i, column=j, value=val)
            wb.save(path)
            self.info_label.config(text=f"💾 Saved: {path.split('/')[-1]}")
            self.save_btn.config(state="disabled")
        except Exception as e:
            messagebox.showerror("Error", f"Cannot save file: {e}")

    def send_matrix(self):
        self.update_matrix_from_entries()
        if not self.matrix:
            messagebox.showwarning("Error", "No matrix to send.")
            return
            
        try:
            response = requests.post(SERVER_UPLOAD, json={"matrix": self.matrix}, timeout=10)
            if response.status_code == 200:
                self.info_label.config(text="🚀 Matrix sent to deciders!")
                self.save_btn.config(state="disabled")
                messagebox.showinfo("Success", "Matrix successfully sent to all deciders!")
            else:
                messagebox.showerror("Server Error", f"{response.status_code}: {response.text}")
        except Exception as e:
            messagebox.showerror("Error", f"Unable to connect to server: {e}")

    def show_deciders_local(self):
        win = tk.Toplevel(self.root)
        win.title("Defined Deciders")
        win.geometry("400x300")
        
        ttk.Label(win, text=f"Expected Deciders ({len(self.deciders_local)} total):", 
                 font=("Arial", 10, "bold")).pack(pady=10)
        
        # Créer le tableau
        cols = ("Decider Name", "Weight (%)")
        tree = ttk.Treeview(win, columns=cols, show="headings", height=8)
        
        tree.heading("Decider Name", text="Decider Name")
        tree.heading("Weight (%)", text="Weight (%)")
        
        tree.column("Decider Name", width=250)
        tree.column("Weight (%)", width=100, anchor="center")
        
        tree.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Remplir avec les données
        total_weight = 0
        for d in self.deciders_local:
            tree.insert("", "end", values=(d["name"], f"{d['weight']:.1f}%"))
            total_weight += d["weight"]
        
        # Afficher le poids total
        ttk.Label(win, text=f"Total Weight: {total_weight:.1f}%", 
                 font=("Arial", 10, "bold")).pack(pady=5)
        
        ttk.Button(win, text="Close", command=win.destroy).pack(pady=10)


if __name__ == "__main__":
    root = tk.Tk()
    app = CoordinatorApp(root)
    root.mainloop()