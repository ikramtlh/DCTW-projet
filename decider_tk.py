import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import requests
import socketio
import threading
from openpyxl import load_workbook, Workbook

# Server URLs
SERVER_UPLOAD = "http://192.168.1.19:5003/upload_matrix"
SERVER_WS = "http://192.168.1.19:5003"


class CoordinatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("DCTW Coordinator")
        self.root.geometry("900x600")

        self.matrix = []
        self.entries = []
        self.deciders_local = [
            {"name": "decider_policeman", "weight": 40.0},
            {"name": "decider_economist", "weight": 25.0},
            {"name": "decider_environmental representative", "weight": 20.0},
            {"name": "decider_public representative", "weight": 15.0},
        ]

        self.sio = None
        self.sio_thread = None
        self.received_rankings = {}

        # Top frame
        top = ttk.Frame(root)
        top.pack(fill="x", padx=10, pady=8)
        self.info_label = ttk.Label(top, text="🔌 Connecting to server...", font=("Arial", 14, "bold"))
        self.info_label.pack(side="left")

        # Buttons
        btn_frame = ttk.Frame(root)
        btn_frame.pack(fill="x", padx=10, pady=6)
        ttk.Button(btn_frame, text="📂 Upload Excel", command=self.load_excel).pack(side="left", padx=4)
        ttk.Button(btn_frame, text="➕ New", command=self.create_matrix_dialog).pack(side="left", padx=4)
        self.save_btn = ttk.Button(btn_frame, text="💾 Save Excel", command=self.save_excel, state="disabled")
        self.save_btn.pack(side="left", padx=4)
        self.send_btn = ttk.Button(btn_frame, text="🚀 Send", command=self.send_matrix, state="disabled")
        self.send_btn.pack(side="left", padx=4)
        ttk.Button(btn_frame, text="👥 Show Deciders", command=self.show_deciders_local).pack(side="left", padx=4)
        self.aggregate_btn = ttk.Button(btn_frame, text="⚙️ Aggregate",
                                        command=self.aggregate_action,
                                        state="disabled")
        self.aggregate_btn.pack(side="left", padx=4)

        # Rankings frame (au-dessus de la matrice)
        self.rankings_frame = ttk.Frame(root)
        self.rankings_frame.pack(fill="x", padx=10, pady=6)

        # Canvas pour la matrice scrollable
        self.canvas = tk.Canvas(root)
        self.frame_container = ttk.Frame(self.canvas)
        self.vsb = ttk.Scrollbar(root, orient="vertical", command=self.canvas.yview)
        self.hsb = ttk.Scrollbar(root, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=self.vsb.set, xscrollcommand=self.hsb.set)
        self.vsb.pack(side="right", fill="y")
        self.hsb.pack(side="bottom", fill="x")
        self.canvas.pack(fill="both", expand=True, padx=10, pady=6)
        self.canvas.create_window((0, 0), window=self.frame_container, anchor="nw")
        self.frame_container.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))

        self.start_socketio_client()

    # ------------------- SOCKET.IO -------------------
    def start_socketio_client(self):
        def run_client():
            try:
                self.sio = socketio.Client(logger=False, reconnection=True)

                @self.sio.on('final_ranking')
                def on_final_ranking(data):
                    # print(f"📊 Received ranking from {data['decider']}: {data['ranking']}")
                    self.received_rankings[data['decider']] = {
                        'phi': data.get('phi', [0]*len(data['ranking'])),
                        'ranking': data['ranking']
                    }
                    # Active le bouton Aggréger si les 4 rankings sont reçus
                    if len(self.received_rankings) == 4:
                      self.root.after(0, lambda: self.aggregate_btn.config(state="normal"))
                    self.root.after(0, self._update_status)

                @self.sio.event
                def connect():
                    print("🔌 Coordinator connected to server for rankings")
                    self.root.after(0, lambda: self.info_label.config(text="✅ Connected - Waiting for rankings"))

                @self.sio.event
                def disconnect():
                    self.root.after(0, lambda: self.info_label.config(text="🔌 Disconnected from server"))

                self.sio.connect(SERVER_WS)
                self.sio.wait()
            except Exception as e:
                self.root.after(0, lambda err=e: self.info_label.config(text=f"❌ SocketIO Error: {str(err)[:50]}"))
 
        self.sio_thread = threading.Thread(target=run_client, daemon=True)
        self.sio_thread.start()

    # ------------------- STATUS & RANKINGS -------------------
    def _update_status(self):
        received_count = len(self.received_rankings)
        total_deciders = len(self.deciders_local)
        status_text = f"📊 {received_count}/{total_deciders} rankings received"
        self.info_label.config(text=status_text)
        self.display_rankings_above_matrix()

    def display_rankings_above_matrix(self):
        # Clear previous
        for widget in self.rankings_frame.winfo_children():
            widget.destroy()

        if not self.received_rankings:
            return

        actions = [row[0] for row in self.matrix[1:]] if self.matrix and len(self.matrix) > 1 else []

        for decider_name, data in self.received_rankings.items():
            ttk.Label(self.rankings_frame, text=f"{decider_name} Ranking:", font=("Arial", 10, "bold")).pack(anchor="w", padx=5, pady=(5,2))

            tree = ttk.Treeview(self.rankings_frame, columns=("Rank", "Action"), show="headings", height=4)
            tree.heading("Rank", text="#")
            tree.heading("Action", text="Action")
            tree.column("Rank", width=50, anchor="center")
            tree.column("Action", width=300)
            tree.pack(fill="x", padx=5, pady=2)

            for pos, action_idx in enumerate(data['ranking']):
                action_name = actions[action_idx] if action_idx < len(actions) else f"Action {action_idx+1}"
                tree.insert("", "end", values=(pos+1, action_name))

    def show_decider_ranking(self, parent_win, decider_name):
        # Clear previous tree if exists
        for widget in parent_win.winfo_children():
            if isinstance(widget, ttk.Treeview):
                widget.destroy()

        actions = [row[0] for row in self.matrix[1:]] if self.matrix and len(self.matrix) > 1 else []
        data = self.received_rankings[decider_name]

        ttk.Label(parent_win, text=f"{decider_name} Ranking:", font=("Arial", 10, "bold")).pack(anchor="w", padx=10, pady=(10,2))

        tree = ttk.Treeview(parent_win, columns=("Rank", "Action"), show="headings", height=35)
        tree.heading("Rank", text="#")
        tree.heading("Action", text="Action")
        tree.column("Rank" )
        tree.column("Action")
        tree.pack(fill="both", expand=False, padx=5, pady=3)

        for pos, action_idx in enumerate(data['ranking']):
            action_name = actions[action_idx] if action_idx < len(actions) else f"Action {action_idx+1}"
            tree.insert("", "end", values=(pos+1, action_name))

    # ------------------- AGGREGATION -------------------
    def aggregate_action(self):
        win = tk.Toplevel(self.root)
        win.title("calculate the score of each action")

        # ---- Taille réduite ----
        win.geometry("240x140")
        win.resizable(False, False)

        # ---- Titre compact ----
        ttk.Label(win, text="calculate the score of each action", font=("Arial", 11, "bold")).pack(pady=10)

        # ---- Bouton Scorage (seul bouton) ----
        score_btn = ttk.Button(win, text="⚙️  Scoring", width=18, command=self.open_scoring_window)
        score_btn.pack(pady=10)

    def open_scoring_window(self):
        # Vérifier que tous les rankings sont reçus
        if len(self.received_rankings) < len(self.deciders_local):
            messagebox.showwarning("Attention", "Tous les rankings des décideurs ne sont pas encore reçus.")
            return

        actions = [row[0] for row in self.matrix[1:]]  # noms des actions
        n_actions = len(actions)

        # Calcul des scores pondérés
        action_scores = {action: 0.0 for action in actions}
        log = ["📄 Start of the calculation of weighted scores :"]

        for decider in self.deciders_local:
            name = decider["name"]
            weight = decider["weight"] / 100.0
            ranking_data = self.received_rankings.get(name)
            if not ranking_data:
                continue
            ranking = ranking_data["ranking"]

            for pos, action_idx in enumerate(ranking):
                action_name = actions[action_idx]
                score = (n_actions - pos) * weight
                action_scores[action_name] += score
                log.append(f"- {name} (poids {weight*100:.1f}%) → {action_name} +{score:.2f}")

        log.append("✅ Scores calculated successfully ")

        # --- Fenêtre principale ---
        win = tk.Toplevel(self.root)
        win.title("Scoring")
        win.geometry("700x400")  # largeur plus grande pour tableau + log

        ttk.Label(win, text=" Action Scoring ", font=("Arial", 12, "bold")).pack(pady=10)

        # Frame principale pour tout
        main_frame = ttk.Frame(win)
        main_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # --- Petit tableau à gauche ---
        frame_table = ttk.Frame(main_frame, width=200)
        frame_table.pack(side="left", fill="y", padx=(0,10))

        tree = ttk.Treeview(frame_table, columns=("Action", "Score"), show="headings", height=15)
        tree.heading("Action", text="Action")
        tree.heading("Score", text="Score")
        tree.column("Action", width=120, anchor="center")
        tree.column("Score", width=50, anchor="center")

        vsb_tree = ttk.Scrollbar(frame_table, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb_tree.set)
        vsb_tree.pack(side="right", fill="y")
        tree.pack(fill="y", expand=True)

        for action, score in sorted(action_scores.items(), key=lambda x: -x[1]):
            tree.insert("", "end", values=(action, f"{score:.2f}"))

        # --- Journal / log à droite ---
        frame_log = ttk.Frame(main_frame)
        frame_log.pack(side="right", fill="both", expand=True)

        ttk.Label(frame_log, text="Negotiation Log", font=("Arial", 10, "bold")).pack(pady=(0,5))

        txt_log = tk.Text(frame_log, wrap="none", height=20)
        txt_log.pack(side="left", fill="both", expand=True)

        vsb_log = ttk.Scrollbar(frame_log, orient="vertical", command=txt_log.yview)
        txt_log.configure(yscrollcommand=vsb_log.set)
        vsb_log.pack(side="right", fill="y")

        for line in log:
          txt_log.insert("end", line + "\n")

        txt_log.config(state="disabled")

        # --- Bouton Démarrer la négociation ---
        def start_negotiation():
            if not tree.get_children():
                messagebox.showwarning("Erreur", "Le tableau des scores est vide !")
                return

            # 1. Récupérer les actions ordonnées par score
            actions_ordered = [tree.item(i)["values"][0] for i in tree.get_children()]
            
            # 2. Récupérer la liste complète des actions depuis la matrice
            all_actions = []
            if hasattr(self, 'matrix') and self.matrix and len(self.matrix) > 1:
                all_actions = [str(row[0]) for row in self.matrix[1:]]
            else:
                all_actions = [str(action) for action in actions_ordered]
            
            # 3. Décideurs
            deciders = list(self.received_rankings.keys())

            txt_log.config(state="normal")
            txt_log.delete("1.0", "end")

            agreement_threshold = 0.9   # 90%
            top_percentage_threshold = 0.7  # 70%
            agreed_action = None
            tour_num = 1

            txt_log.insert("end", "=== DÉBUT DE LA NÉGOCIATION ===\n\n")
            
            # 4. AFFICHER POUR DÉBUG (important!)
            txt_log.insert("end", "=== INFORMATION DE DÉBUG ===\n")
            txt_log.insert("end", f"Actions ordonnées par score: {actions_ordered}\n")
            txt_log.insert("end", f"Toutes les actions: {all_actions}\n")
            
            for d in deciders:
                ranking_data = self.received_rankings[d]
                ranking = ranking_data["ranking"]
                txt_log.insert("end", f"\n{d}: ranking = {ranking[:5]}... (total: {len(ranking)})\n")
                
                # Essayer de déterminer si ce sont des indices ou des numéros
                if ranking:
                    first = str(ranking[0])
                    if first.isdigit() and len(first) >= 4:  # C'est probablement un numéro d'action
                        txt_log.insert("end", f"  → Format: NUMÉROS D'ACTION\n")
                    else:
                        txt_log.insert("end", f"  → Format: INDICES (0-{len(all_actions)-1})\n")
            
            txt_log.insert("end", "\n" + "="*40 + "\n\n")
            
            # 5. PROCESSUS DE NÉGOCIATION
            for action in actions_ordered:
                action_str = str(action)
                responses = {}
                
                for d in deciders:
                    ranking_data = self.received_rankings[d]
                    ranking = ranking_data["ranking"]
                    
                    total = len(ranking)
                    top_n = int(total * top_percentage_threshold)
                    
                    # Vérifier le format des données
                    if ranking:
                        first = str(ranking[0])
                        # Si c'est un nombre de 4 chiffres ou plus → c'est un numéro d'action
                        if first.isdigit() and len(first) >= 4:
                            # FORMAT 1: ranking contient des numéros d'action
                            top_list = [str(item) for item in ranking[:top_n]]
                            if action_str in top_list:
                                responses[d] = "oui"
                            else:
                                responses[d] = "non"
                        else:
                            # FORMAT 2: ranking contient des indices
                            try:
                                # Trouver l'index de l'action dans all_actions
                                action_index = all_actions.index(action_str)
                            except ValueError:
                                action_index = -1
                            
                            if action_index != -1 and action_index in ranking[:top_n]:
                                responses[d] = "oui"
                            else:
                                responses[d] = "non"
                    else:
                        responses[d] = "non"  # Si ranking vide

                # Calcul du pourcentage
                oui_count = list(responses.values()).count("oui")
                agreement_ratio = oui_count / len(deciders)

                # Affichage dans le journal
                txt_log.insert("end", f"Tour {tour_num} - Proposition: {action_str}\n")
                for d in deciders:
                    r = responses.get(d, "non")
                    txt_log.insert("end", f"  {d} [{r}]\n")
                
                txt_log.insert("end", f"  → Accord actuel: {agreement_ratio*100:.1f}%\n\n")
                txt_log.see("end")
                txt_log.update()

                # Vérifier si seuil atteint
                if agreement_ratio >= agreement_threshold:
                    agreed_action = action_str
                    break

                tour_num += 1

            # 6. RÉSULTAT FINAL
            txt_log.insert("end", "="*50 + "\n")
            txt_log.insert("end", "RÉSULTAT FINAL DE LA NÉGOCIATION\n")
            txt_log.insert("end", "="*50 + "\n\n")
            
            if agreed_action:
                txt_log.insert("end", f"✅ ACTION RETENUE: {agreed_action}\n\n")
                
                # Afficher les votes finaux
                txt_log.insert("end", "Votes finaux:\n")
                for d in deciders:
                    ranking = self.received_rankings[d]["ranking"]
                    total = len(ranking)
                    top_n = int(total * top_percentage_threshold)
                    
                    if str(ranking[0]).isdigit() and len(str(ranking[0])) >= 4:
                        top_list = [str(item) for item in ranking[:top_n]]
                        vote = "oui" if agreed_action in top_list else "non"
                    else:
                        action_index = all_actions.index(agreed_action) if agreed_action in all_actions else -1
                        vote = "oui" if action_index != -1 and action_index in ranking[:top_n] else "non"
                    
                    txt_log.insert("end", f"  {d}: {vote}\n")
            else:
                txt_log.insert("end", "⚠️ AUCUNE ACTION N'A ATTEINT L'ACCORD REQUIS (90%)\n\n")
                
                # Afficher le top 3 avec leurs pourcentages
                txt_log.insert("end", "Meilleures propositions examinées:\n")
                for i in range(min(3, len(actions_ordered))):
                    action = actions_ordered[i]
                    scores = []
                    for d in deciders:
                        ranking = self.received_rankings[d]["ranking"]
                        total = len(ranking)
                        top_n = int(total * top_percentage_threshold)
                        
                        if ranking and str(ranking[0]).isdigit() and len(str(ranking[0])) >= 4:
                            top_list = [str(item) for item in ranking[:top_n]]
                            scores.append(1 if str(action) in top_list else 0)
                        else:
                            action_index = all_actions.index(str(action)) if str(action) in all_actions else -1
                            scores.append(1 if action_index != -1 and action_index in ranking[:top_n] else 0)
                    
                    agreement = sum(scores) / len(scores)
                    txt_log.insert("end", f"  {i+1}. {action}: {agreement*100:.1f}%\n")

            txt_log.see("end")
            txt_log.config(state="disabled")


        # Ajouter le bouton juste sous le tableau
        start_btn = ttk.Button(frame_table, text="▶ Start the Negotiation ", command=start_negotiation)
        start_btn.pack(pady=15)

        # Journal plus petit et vide au départ
        txt_log.config(height=10)  # hauteur réduite
        txt_log.delete("1.0", "end")
        txt_log.config(state="normal")

    # ------------------- MATRIX GRID -------------------
    def clear_grid(self):
        for row in self.entries:
            for e in row:
                e.destroy()
        self.entries = []

    def build_grid(self):
        self.clear_grid()
        rows = len(self.matrix)
        cols = len(self.matrix[0]) if rows > 0 else 0
        self.entries = [[None for _ in range(cols)] for _ in range(rows)]
        for i in range(rows):
            for j in range(cols):
                val = self.matrix[i][j] if i < len(self.matrix) and j < len(self.matrix[i]) else ""
                e = ttk.Entry(self.frame_container, width=15)
                e.grid(row=i, column=j, padx=2, pady=2, sticky="nsew")
                e.insert(0, val)
                e.bind("<KeyRelease>", lambda ev: self.on_edit())
                self.entries[i][j] = e

    def update_matrix_from_entries(self):
        self.matrix = [[cell.get() for cell in row] for row in self.entries]

    def on_edit(self):
        self.save_btn.config(state="normal")
        self.send_btn.config(state="normal")

    # ------------------- EXCEL -------------------
    def load_excel(self):
        path = filedialog.askopenfilename(title="Select Excel file", filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            wb = load_workbook(path)
            ws = wb.active
            self.matrix = [[str(c) if c is not None else "" for c in r] for r in ws.iter_rows(values_only=True)]
            if not self.matrix:
                messagebox.showwarning("Empty", "The Excel sheet is empty.")
                return
            self.build_grid()
            self.save_btn.config(state="normal")
            self.send_btn.config(state="normal")
            self.info_label.config(text=f"📂 Loaded: {path.split('/')[-1]}")
        except Exception as e:
            messagebox.showerror("Error", f"Cannot load Excel file: {e}")

    def create_matrix_dialog(self):
        rows = simpledialog.askinteger("Rows", "How many choices (rows)?", initialvalue=3, minvalue=1, maxvalue=50)
        if rows is None:
            return
        cols = simpledialog.askinteger("Columns", "How many criteria (columns)?", initialvalue=3, minvalue=1, maxvalue=50)
        if cols is None:
            return
        self.matrix = [["" for _ in range(cols)] for _ in range(rows)]
        self.build_grid()
        self.save_btn.config(state="normal")
        self.send_btn.config(state="normal")
        self.info_label.config(text="➕ New Matrix")

    def save_excel(self):
        self.update_matrix_from_entries()
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            for i, row in enumerate(self.matrix, start=1):
                for j, val in enumerate(row, start=1):
                    ws.cell(row=i, column=j, value=val)
            wb.save(path)
            self.info_label.config(text=f"💾 Saved: {path.split('/')[-1]}")
            self.save_btn.config(state="disabled")
        except Exception as e:
            messagebox.showerror("Error", f"Cannot save file: {e}")

    def send_matrix(self):
        self.update_matrix_from_entries()
        if not self.matrix:
            messagebox.showwarning("Error", "No matrix to send.")
            return
        try:
            r = requests.post(SERVER_UPLOAD, json={"matrix": self.matrix}, timeout=5)
            if r.status_code == 200:
                self.info_label.config(text="🚀 Matrix sent to deciders!")
                self.save_btn.config(state="disabled")
            else:
                messagebox.showerror("Server Error", f"{r.status_code}: {r.text}")
        except Exception as e:
            messagebox.showerror("Error", f"Unable to connect to server: {e}")

    def show_deciders_local(self):
        win = tk.Toplevel(self.root)
        win.title("Defined Deciders")
        win.geometry("450x350")
        ttk.Label(win, text=f"Expected Deciders ({len(self.deciders_local)} total):", font=("Arial", 10, "bold")).pack(pady=10)
        cols = ("Decider Name", "Weight (%)")
        tree = ttk.Treeview(win, columns=cols, show="headings", height=8)
        tree.heading("Decider Name", text="Decider Name")
        tree.heading("Weight (%)", text="Weight (%)")
        tree.column("Decider Name", width=250)
        tree.column("Weight (%)", width=100, anchor="center")
        tree.pack(fill="both", expand=True, padx=10, pady=10)
        total_weight = 0
        for d in self.deciders_local:
            tree.insert("", "end", values=(d["name"], f"{d['weight']:.1f}%"))
            total_weight += d["weight"]
        ttk.Label(win, text=f"Total Weight: {total_weight:.1f}%", font=("Arial", 10, "bold")).pack(pady=5)
        ttk.Button(win, text="Close", command=win.destroy).pack(pady=10)



if __name__ == "__main__":
    root = tk.Tk()
    app = CoordinatorApp(root)
    root.mainloop()