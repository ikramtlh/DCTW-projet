import sys
import requests
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton,
    QLabel, QTextEdit, QFileDialog, QHBoxLayout,
    QTableWidget, QTableWidgetItem, QInputDialog
)
from PySide6.QtGui import QTextCursor, QTextBlockFormat, Qt
from openpyxl import Workbook, load_workbook  # pour Excel

SERVER_URL = "http://localhost:5002"

class CoordinatorWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Coordinateur DCTW")
        self.setGeometry(600, 300, 800, 600)

        self.matrix = None
        self.deciders_data = None
        self.is_modified = False  # Nouveau flag pour modifications

        # Layout principal
        layout = QVBoxLayout()

        # Label info
        self.info_label = QLabel("Welcome coordinateur")
        layout.addWidget(self.info_label)

        # Zone de texte pour messages
        self.deciders_text = QTextEdit()
        self.deciders_text.setReadOnly(True)
        layout.addWidget(self.deciders_text)

        # Table pour afficher les décideurs
        self.decider_table = QTableWidget()
        self.decider_table.setColumnCount(2)
        self.decider_table.setHorizontalHeaderLabels(["Décideur", "Poids (%)"])
        self.decider_table.horizontalHeader().setStretchLastSection(True)
        self.decider_table.setVisible(False)
        layout.addWidget(self.decider_table)

        # Table pour la matrice remplissable
        self.matrix_table = QTableWidget()
        self.matrix_table.setVisible(False)
        self.matrix_table.cellChanged.connect(self.on_matrix_modified)  # Détection modification
        layout.addWidget(self.matrix_table)

        # Layout pour les boutons
        btn_layout = QHBoxLayout()

        self.upload_btn = QPushButton("Upload")
        self.upload_btn.clicked.connect(self.upload_matrix)
        btn_layout.addWidget(self.upload_btn)

        self.new_btn = QPushButton("New")
        self.new_btn.clicked.connect(self.new_matrix)
        btn_layout.addWidget(self.new_btn)

        self.save_btn = QPushButton("Save")
        self.save_btn.setEnabled(False)  # Désactivé initialement
        self.save_btn.clicked.connect(self.save_matrix)
        btn_layout.addWidget(self.save_btn)

        self.send_btn = QPushButton("Send")
        self.send_btn.setEnabled(False)
        self.send_btn.clicked.connect(self.send_matrix)
        btn_layout.addWidget(self.send_btn)

        self.decider_btn = QPushButton("Decider")
        self.decider_btn.clicked.connect(self.show_deciders_table)
        btn_layout.addWidget(self.decider_btn)

        layout.addLayout(btn_layout)
        self.setLayout(layout)

        # Charger les décideurs au démarrage
        self.load_deciders()

    def center_text(self, text):
        """Permet de centrer le texte dans QTextEdit"""
        self.deciders_text.setText(text)
        cursor = self.deciders_text.textCursor()
        block_format = QTextBlockFormat()
        block_format.setAlignment(Qt.AlignmentFlag.AlignCenter)
        cursor.select(QTextCursor.SelectionType.Document)
        cursor.setBlockFormat(block_format)
        cursor.clearSelection()
        self.deciders_text.setTextCursor(cursor)

    def load_deciders(self):
        """Charger la liste des décideurs et poids depuis le serveur"""
        try:
            r = requests.get(SERVER_URL)
            data = r.json()
            self.deciders_data = data
        except Exception as e:
            self.center_text(f"Erreur serveur : {e}")

    def show_deciders_table(self):
        """Afficher les décideurs et leurs poids dans un tableau"""
        if not self.deciders_data:
            self.center_text("Pas de données des décideurs")
            return
        deciders = self.deciders_data.get("deciders", [])
        weights = self.deciders_data.get("weights", {})

        self.decider_table.setRowCount(len(deciders))
        for i, d in enumerate(deciders):
            item_decider = QTableWidgetItem(d)
            item_decider.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.decider_table.setItem(i, 0, item_decider)

            item_weight = QTableWidgetItem(str(weights[d]))
            item_weight.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.decider_table.setItem(i, 1, item_weight)

        self.decider_table.setVisible(True)

    def on_matrix_modified(self, row, column):
        """Active le bouton Save si la matrice est modifiée"""
        if not self.save_btn.isEnabled():
            self.save_btn.setEnabled(True)
        self.is_modified = True

    def upload_matrix(self):
        """Uploader une matrice depuis un fichier Excel"""
        options = QFileDialog.Options()
        filename, _ = QFileDialog.getOpenFileName(
            self,
            "Choisir fichier Excel",
            "",
            "Excel Files (*.xlsx);;All Files (*)",
            options=options
        )
        if filename:
            try:
                wb = load_workbook(filename)
                ws = wb.active

                # Extraire les données pour la matrice
                self.matrix = []
                for row in ws.iter_rows(values_only=True):
                    self.matrix.append([str(cell) if cell is not None else "" for cell in row])

                # Afficher dans la table
                self.display_matrix_table(self.matrix)
                self.info_label.setText(f"✅ Matrice chargée depuis Excel : {filename}")
                self.send_btn.setEnabled(True)
                self.save_btn.setEnabled(False)  # Save désactivé jusqu'à modification
                self.is_modified = False
            except Exception as e:
                self.info_label.setText(f"Erreur lecture fichier Excel : {e}")

    def new_matrix(self):
        """Créer une nouvelle matrice interactive"""
        rows, ok1 = QInputDialog.getInt(self, "Nombre de choix", "Combien de choix (lignes) ?", 3, 1, 20, 1)
        if not ok1:
            return
        cols, ok2 = QInputDialog.getInt(self, "Nombre de critères", "Combien de critères (colonnes) ?", 3, 1, 20, 1)
        if not ok2:
            return

        self.matrix_table.setRowCount(rows)
        self.matrix_table.setColumnCount(cols)
        self.matrix_table.setHorizontalHeaderLabels([f"C{i+1}" for i in range(cols)])
        self.matrix_table.setVerticalHeaderLabels([f"L{i+1}" for i in range(rows)])
        self.matrix_table.setVisible(True)
        self.center_text("✅ Nouvelle matrice créée ! Remplissez la table ci-dessus.")
        self.save_btn.setEnabled(False)
        self.send_btn.setEnabled(True)
        self.is_modified = False

    def save_matrix(self):
        """Récupérer la matrice depuis la table, afficher et sauvegarder Excel"""
        if not self.matrix_table.isVisible():
            self.center_text("Pas de matrice à sauvegarder")
            return

        rows = self.matrix_table.rowCount()
        cols = self.matrix_table.columnCount()
        self.matrix = []
        for i in range(rows):
            row_data = []
            for j in range(cols):
                item = self.matrix_table.item(i, j)
                value = item.text() if item else ""
                row_data.append(value)
            self.matrix.append(row_data)

        # Afficher la matrice dans la table
        self.display_matrix_table(self.matrix)
        self.info_label.setText("✅ Matrice sauvegardée :")

        # Sauvegarder dans Excel
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Enregistrer matrice comme Excel",
            "",
            "Excel Files (*.xlsx)",
            options=options
        )
        if file_path:
            if not file_path.endswith(".xlsx"):
                file_path += ".xlsx"
            try:
                wb = Workbook()
                ws = wb.active
                for i, row in enumerate(self.matrix, start=1):
                    for j, value in enumerate(row, start=1):
                        ws.cell(row=i, column=j, value=value)
                wb.save(file_path)
                self.info_label.setText(f"✅ Matrice sauvegardée en Excel : {file_path}")
                self.save_btn.setEnabled(False)  # Save désactivé après sauvegarde
                self.is_modified = False
            except Exception as e:
                self.info_label.setText(f"Erreur sauvegarde Excel: {e}")

    def display_matrix_table(self, matrix):
        """Afficher une matrice existante dans la table"""
        rows = len(matrix)
        cols = len(matrix[0]) if rows > 0 else 0
        self.matrix_table.blockSignals(True)  # Empêche l'activation de cellChanged pendant remplissage
        self.matrix_table.setRowCount(rows)
        self.matrix_table.setColumnCount(cols)
        self.matrix_table.setHorizontalHeaderLabels([f"C{i+1}" for i in range(cols)])
        self.matrix_table.setVerticalHeaderLabels([f"L{i+1}" for i in range(rows)])
        for i in range(rows):
            for j in range(cols):
                item = QTableWidgetItem(str(matrix[i][j]))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.matrix_table.setItem(i, j, item)
        self.matrix_table.blockSignals(False)
        self.matrix_table.setVisible(True)

    def send_matrix(self):
        """Envoyer la matrice remplie au serveur"""
        if not self.matrix:
            self.center_text("Aucune matrice à envoyer !")
            return
        try:
            r = requests.post(f"{SERVER_URL}/upload_matrix", json={"matrix": self.matrix})
            self.center_text("✅ Matrice envoyée au serveur !")
            print("Réponse serveur:", r.json())
        except Exception as e:
            self.center_text(f"Erreur envoi serveur : {e}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = CoordinatorWindow()
    window.show()
    sys.exit(app.exec())
